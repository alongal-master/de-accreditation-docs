#!/usr/bin/env python3
"""
Syllabus Generator for Four-Week Course

This script creates an Excel syllabus file from a list of lesson titles.
It organizes lessons into 4 weeks with 3-4 chapters each, generates learning goals,
and estimates completion times to total 30 hours per week.
"""

import argparse
import os
import sys
import json
from typing import List, Dict, Tuple
import openai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from prompts import (
    get_lesson_parsing_prompt, 
    get_week_learning_goals_prompt, 
    get_chapter_info_prompt,
    get_practice_sessions_for_week_prompt,
    get_maestro_json_prompt
)

# Course structure constants
NUM_OF_WEEKS = 4
NUM_OF_CHAPTERS_PER_WEEK = 5
NUM_PRACTICE_SESSIONS_PER_CHAPTER = 1
NUM_LIVE_SESSIONS_PER_CHAPTER = 1
LIVE_SESSION_TITLE = "Sync Session: Q&A"
NUM_WEEKLY_REVIEW_PER_WEEK = 1
WEEKLY_REVIEW_TITLE = "Weekly Review"

# Base time estimates (in minutes) for different lesson types
BASE_TIME_ESTIMATES = {
    "practice": 60,
    "weekly review": 120,
    "final assessment": 120,
    "sync session": 90,
    "default": 45
}

# Export options
EXPORT_TO_LESSONS_JSON = False
EXPORT_TO_MAESTRO_JSON = False

# AI Model
#DEFAULT_MODEL = "gpt-4-turbo"
#DEFAULT_MODEL = "gpt-4o-mini"
DEFAULT_MODEL = "gpt-5.1"

# Pricing per 1M tokens (as of 2024, update as needed)
MODEL_PRICING = {
    "gpt-5.1": {"input": 2.50, "output": 10.00},  # Update with actual pricing when available
    "gpt-4o": {"input": 2.50, "output": 10.00},
    "gpt-4o-mini": {"input": 0.15, "output": 0.60},
    "gpt-4-turbo": {"input": 10.00, "output": 30.00},
    "gpt-4": {"input": 30.00, "output": 60.00},
    "gpt-3.5-turbo": {"input": 0.50, "output": 1.50},
}

# Load environment variables
load_dotenv()

class SyllabusGenerator:
    def __init__(self, api_key: str, log_file: str = "openai_log.txt", model: str = DEFAULT_MODEL):
        """Initialize the syllabus generator with OpenAI API key."""
        self.api_key = api_key
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Course Syllabus"
        self.log_file = log_file
        self.model = model
        self._init_log_file()
    
    def _init_log_file(self):
        """Initialize the log file with header."""
        with open(self.log_file, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("OpenAI API Log - Syllabus Generator\n")
            f.write("=" * 80 + "\n")
            f.write(f"Started at: {self._get_timestamp()}\n")
            f.write(f"Model: {self.model}\n")
            f.write("=" * 80 + "\n\n")
    
    def _get_timestamp(self):
        """Get current timestamp for logging."""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def _log_api_call(self, prompt: str, response: str, call_type: str):
        """Log API call details to file."""
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{self._get_timestamp()}] {call_type}\n")
            f.write("-" * 60 + "\n")
            f.write("PROMPT:\n")
            f.write(prompt)
            f.write("\n\n")
            f.write("RESPONSE:\n")
            f.write(response)
            f.write("\n\n")
            f.write("=" * 80 + "\n\n")
    
    def _get_openai_client(self):
        """Get OpenAI client."""
        return openai.OpenAI(api_key=self.api_key)
    
    def _get_max_tokens_param(self, max_tokens_value: int) -> dict:
        """Get the correct max tokens parameter based on model version.
        
        gpt-5.1 uses max_completion_tokens, older models use max_tokens.
        """
        if self.model.startswith("gpt-5"):
            return {"max_completion_tokens": max_tokens_value}
        else:
            return {"max_tokens": max_tokens_value}
    
    def _calculate_cost(self, input_tokens: int, output_tokens: int) -> float:
        """Calculate the cost of an API call based on token usage."""
        if self.model not in MODEL_PRICING:
            print(f"Warning: Pricing not available for model {self.model}")
            return 0.0
        
        pricing = MODEL_PRICING[self.model]
        input_cost = (input_tokens / 1_000_000) * pricing["input"]
        output_cost = (output_tokens / 1_000_000) * pricing["output"]
        total_cost = input_cost + output_cost
        return total_cost
    
    def read_lessons_from_file(self, file_path: str) -> List[Dict[str, str]]:
        """Read and parse lesson titles and learning outcomes from a text file using AI."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            if not content:
                print("Error: File is empty.")
                sys.exit(1)
            
            # Use AI to parse the content into structured lesson data
            return self._parse_lessons_with_ai(content)
            
        except FileNotFoundError:
            print(f"Error: File '{file_path}' not found.")
            sys.exit(1)
        except Exception as e:
            print(f"Error reading file: {e}")
            sys.exit(1)
    
    def _parse_lessons_with_ai(self, content: str) -> List[Dict[str, str]]:
        """Use AI to parse free text lessons and learning outcomes into structured data."""
        prompt = get_lesson_parsing_prompt(content)
        
        try:
            client = self._get_openai_client()
            print("Parsing lessons with AI (streaming)...")
            
            # Use streaming for chunked responses
            stream = client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                **self._get_max_tokens_param(4000),
                temperature=0.1,
                stream=True,
                stream_options={"include_usage": True}
            )
            
            # Collect chunks and track usage
            ai_content = ""
            usage_data = None
            for chunk in stream:
                # Skip chunks without choices
                if not chunk.choices or len(chunk.choices) == 0:
                    continue
                    
                if chunk.choices[0].delta.content is not None:
                    content_chunk = chunk.choices[0].delta.content
                    ai_content += content_chunk
                    # Show progress
                    if len(ai_content) % 100 == 0:
                        print(".", end="", flush=True)
                # Capture usage data from the final chunk
                if hasattr(chunk, 'usage') and chunk.usage is not None:
                    usage_data = chunk.usage
            
            print()  # New line after progress dots
            
            # Display cost information
            if usage_data:
                cost = self._calculate_cost(usage_data.prompt_tokens, usage_data.completion_tokens)
                print(f"📊 Tokens: {usage_data.prompt_tokens:,} in + {usage_data.completion_tokens:,} out = {usage_data.total_tokens:,} total | 💰 ${cost:.4f}")
            
            # Log the API call
            self._log_api_call(prompt, ai_content, "LESSON PARSING (STREAMED)")
            
            lessons = []
            
            # Parse the structured response
            lines = ai_content.split('\n')
            current_lesson = {}
            collecting_outcomes = False
            outcomes_lines = []
            
            for line in lines:
                line = line.strip()
                if line.startswith('TITLE:'):
                    # Save previous lesson if exists
                    if current_lesson:
                        # Join any collected outcome lines with semicolon
                        if outcomes_lines:
                            current_lesson['learning_outcomes'] = '; '.join(outcomes_lines)
                        lessons.append(current_lesson)
                    current_lesson = {'title': line[6:].strip(), 'learning_outcomes': ''}
                    collecting_outcomes = False
                    outcomes_lines = []
                elif line.startswith('OUTCOMES:'):
                    if current_lesson:
                        outcomes_text = line[9:].strip()
                        if outcomes_text:
                            outcomes_lines.append(outcomes_text)
                        collecting_outcomes = True
                elif collecting_outcomes and line and not line.startswith('TITLE:'):
                    # Collect additional outcome lines until next TITLE
                    outcomes_lines.append(line)
            
            # Add the last lesson
            if current_lesson and current_lesson.get('title'):
                if outcomes_lines:
                    current_lesson['learning_outcomes'] = '; '.join(outcomes_lines)
                lessons.append(current_lesson)
            
            if not lessons:
                raise ValueError("No lessons found in AI response")
            
            print(f"Successfully parsed {len(lessons)} lessons with AI")
            return lessons
            
        except Exception as e:
            print(f"Error parsing lessons with AI: {e}")
            raise RuntimeError(f"Failed to parse lessons with AI: {e}")
    
    
    def distribute_lessons(self, lessons: List[Dict[str, str]]) -> Dict[int, List[List[Dict[str, str]]]]:
        """
        Distribute lessons across all weeks and chapters.
        Each chapter should have a clear topic.
        """
        total_lessons = len(lessons)
        lessons_per_week = total_lessons // NUM_OF_WEEKS
        remainder = total_lessons % NUM_OF_WEEKS
        
        distribution = {}
        lesson_index = 0
        
        for week in range(1, NUM_OF_WEEKS+1):
            # Distribute remainder lessons to first few weeks
            week_lesson_count = lessons_per_week + (1 if week <= remainder else 0)
            
            lessons_per_chapter = week_lesson_count // NUM_OF_CHAPTERS_PER_WEEK
            chapter_remainder = week_lesson_count % NUM_OF_CHAPTERS_PER_WEEK
            
            week_chapters = []
            for chapter in range(NUM_OF_CHAPTERS_PER_WEEK):
                # Distribute remainder lessons to first few chapters
                chapter_lesson_count = lessons_per_chapter + (1 if chapter < chapter_remainder else 0)
                
                chapter_lessons = lessons[lesson_index:lesson_index + chapter_lesson_count]
                week_chapters.append(chapter_lessons)
                lesson_index += chapter_lesson_count
            
            distribution[week] = week_chapters
        
        return distribution
    
    def add_practice_sessions_to_week(self, week_lessons: List[List[Dict[str, str]]], week_num: int, num_practice_sessions: int) -> List[List[Dict[str, str]]]:
        """Add practice sessions to each chapter in a week using AI."""
        week_lessons_with_practice = []
        
        # Generate practice sessions for ALL chapters in ONE API call
        all_practice_sessions = self.generate_practice_sessions_for_week(week_lessons, week_num, num_practice_sessions)
        
        for chapter_num, (chapter_lessons, practice_sessions) in enumerate(zip(week_lessons, all_practice_sessions), 1):
            # Add original lessons
            chapter_with_practice = chapter_lessons.copy()
            
            # Add practice sessions to the chapter (already generated above)
            chapter_with_practice.extend(practice_sessions)
            
            # Add live sessions to the chapter
            for _ in range(NUM_LIVE_SESSIONS_PER_CHAPTER):
                live_session = {
                    'title': LIVE_SESSION_TITLE,
                    'learning_outcomes': 'Ask questions, clarify concepts, and engage with instructors and peers in real-time.'
                }
                chapter_with_practice = [live_session] + chapter_with_practice  # add to the start
            
            week_lessons_with_practice.append(chapter_with_practice)
        
        # Add weekly reviews to the end of the last chapter
        if week_lessons_with_practice:
            for _ in range(NUM_WEEKLY_REVIEW_PER_WEEK):
                weekly_review = {
                    'title': WEEKLY_REVIEW_TITLE,
                    'learning_outcomes': 'Review the concepts learned in this week and complete weekly assessment.'
                }
                week_lessons_with_practice[-1].append(weekly_review)
        
        return week_lessons_with_practice
    
    def generate_practice_sessions_for_week(self, week_lessons: List[List[Dict[str, str]]], week_num: int, num_sessions_per_chapter: int) -> List[List[Dict[str, str]]]:
        """Generate practice sessions for all chapters in a week using ONE AI call."""
        # Build prompt with all chapters
        chapters_text = ""
        for chapter_num, chapter_lessons in enumerate(week_lessons, 1):
            lessons_text = "\n".join([f"  - {lesson['title']}: {lesson['learning_outcomes']}" for lesson in chapter_lessons])
            chapters_text += f"\nCHAPTER {chapter_num}:\n{lessons_text}\n"
        
        total_sessions = len(week_lessons) * num_sessions_per_chapter
        
        prompt = get_practice_sessions_for_week_prompt(
            chapters_text, 
            week_num, 
            len(week_lessons), 
            num_sessions_per_chapter, 
            total_sessions
        )
        
        try:
            client = self._get_openai_client()
            print(f"Generating practice sessions for Week {week_num} ({len(week_lessons)} chapters, {num_sessions_per_chapter} sessions per chapter)...")
            
            # Use streaming for real-time feedback
            stream = client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                **self._get_max_tokens_param(2000),  # Increased for multiple chapters
                temperature=0.7,
                stream=True,
                stream_options={"include_usage": True}
            )
            
            # Collect chunks and track usage
            ai_content = ""
            usage_data = None
            for chunk in stream:
                # Skip chunks without choices
                if not chunk.choices or len(chunk.choices) == 0:
                    continue
                    
                if chunk.choices[0].delta.content is not None:
                    content_chunk = chunk.choices[0].delta.content
                    ai_content += content_chunk
                    # Show progress
                    if len(ai_content) % 100 == 0:
                        print(".", end="", flush=True)
                # Capture usage data from the final chunk
                if hasattr(chunk, 'usage') and chunk.usage is not None:
                    usage_data = chunk.usage
            
            print()  # New line after progress dots
            
            # Display cost information
            if usage_data:
                cost = self._calculate_cost(usage_data.prompt_tokens, usage_data.completion_tokens)
                print(f"📊 Tokens: {usage_data.prompt_tokens:,} in + {usage_data.completion_tokens:,} out = {usage_data.total_tokens:,} total | 💰 ${cost:.4f}")
            
            # Log the API call
            self._log_api_call(prompt, ai_content, f"WEEK {week_num} ALL CHAPTERS PRACTICE SESSIONS (STREAMED)")
            
            # Parse the structured response grouped by chapter
            all_practice_sessions = []
            current_chapter_sessions = []
            current_session = {}
            
            lines = ai_content.split('\n')
            for line in lines:
                line = line.strip()
                if line.startswith('CHAPTER:'):
                    # Save previous chapter's sessions if any
                    if current_session and current_session.get('title'):
                        current_chapter_sessions.append(current_session)
                        current_session = {}
                    if current_chapter_sessions:
                        all_practice_sessions.append(current_chapter_sessions)
                        current_chapter_sessions = []
                elif line.startswith('TITLE:'):
                    if current_session and current_session.get('title'):
                        current_chapter_sessions.append(current_session)
                    current_session = {'title': line[6:].strip(), 'learning_outcomes': ''}
                elif line.startswith('OUTCOMES:'):
                    if current_session:
                        current_session['learning_outcomes'] = line[9:].strip()
            
            # Add the last session and chapter
            if current_session and current_session.get('title'):
                current_chapter_sessions.append(current_session)
            if current_chapter_sessions:
                all_practice_sessions.append(current_chapter_sessions)
            
            # Validate we have the right number of chapters
            if len(all_practice_sessions) != len(week_lessons):
                raise ValueError(f"Mismatch in chapter count: expected {len(week_lessons)}, got {len(all_practice_sessions)}")
            
            print(f"Generated {sum(len(sessions) for sessions in all_practice_sessions)} practice sessions across {len(all_practice_sessions)} chapters")
            return all_practice_sessions
            
        except Exception as e:
            print(f"Error generating practice sessions for week: {e}")
            raise RuntimeError(f"Failed to generate practice sessions for week {week_num}: {e}")
    
    def generate_learning_goals(self, week_lessons: List[List[Dict[str, str]]], week_num: int, num_practice_sessions: int = 1) -> Tuple[str, List[List[Dict[str, str]]]]:
        """Generate learning goals for a week using OpenAI and return lessons with practice sessions."""
        # Add practice sessions to each chapter
        week_lessons_with_practice = self.add_practice_sessions_to_week(week_lessons, week_num, num_practice_sessions)
        
        all_lessons = [lesson for chapter in week_lessons_with_practice for lesson in chapter]
        lessons_text = "\n".join([f"- {lesson['title']}: {lesson['learning_outcomes']}" for lesson in all_lessons])
        
        prompt = get_week_learning_goals_prompt(lessons_text, week_num)
        
        try:
            client = self._get_openai_client()
            
            # Use streaming for real-time feedback
            stream = client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                **self._get_max_tokens_param(200),
                temperature=0.7,
                stream=True,
                stream_options={"include_usage": True}
            )
            
            # Collect chunks and track usage
            result = ""
            usage_data = None
            for chunk in stream:
                # Skip chunks without choices
                if not chunk.choices or len(chunk.choices) == 0:
                    continue
                    
                if chunk.choices[0].delta.content is not None:
                    result += chunk.choices[0].delta.content
                # Capture usage data from the final chunk
                if hasattr(chunk, 'usage') and chunk.usage is not None:
                    usage_data = chunk.usage
            
            result = result.strip()
            
            # Display cost information
            if usage_data:
                cost = self._calculate_cost(usage_data.prompt_tokens, usage_data.completion_tokens)
                print(f"📊 Tokens: {usage_data.prompt_tokens:,} in + {usage_data.completion_tokens:,} out = {usage_data.total_tokens:,} total | 💰 ${cost:.4f}")
            
            # Log the API call
            self._log_api_call(prompt, result, f"WEEK {week_num} LEARNING GOALS (STREAMED)")
            
            return result, week_lessons_with_practice
        except Exception as e:
            print(f"Error generating learning goals: {e}")
            raise RuntimeError(f"Failed to generate learning goals for week {week_num}: {e}")
    
    def generate_chapter_info(self, chapter_lessons: List[Dict[str, str]], week_num: int, chapter_num: int) -> Tuple[str, str]:
        """Generate chapter title and learning goals using OpenAI."""
        lessons_text = "\n".join([f"- {lesson['title']}: {lesson['learning_outcomes']}" for lesson in chapter_lessons])
        
        prompt = get_chapter_info_prompt(lessons_text)
        
        try:
            client = self._get_openai_client()
            
            # Use streaming for real-time feedback
            stream = client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                **self._get_max_tokens_param(150),
                temperature=0.7,
                stream=True,
                stream_options={"include_usage": True}
            )
            
            # Collect chunks and track usage
            content = ""
            usage_data = None
            for chunk in stream:
                # Skip chunks without choices
                if not chunk.choices or len(chunk.choices) == 0:
                    continue
                    
                if chunk.choices[0].delta.content is not None:
                    content += chunk.choices[0].delta.content
                # Capture usage data from the final chunk
                if hasattr(chunk, 'usage') and chunk.usage is not None:
                    usage_data = chunk.usage
            
            content = content.strip()
            
            # Display cost information
            if usage_data:
                cost = self._calculate_cost(usage_data.prompt_tokens, usage_data.completion_tokens)
                print(f"📊 Tokens: {usage_data.prompt_tokens:,} in + {usage_data.completion_tokens:,} out = {usage_data.total_tokens:,} total | 💰 ${cost:.4f}")
            
            # Log the API call
            self._log_api_call(prompt, content, f"WEEK {week_num} CHAPTER {chapter_num} INFO (STREAMED)")
            
            lines = content.split('\n')
            title = lines[0].replace('TITLE:', '').strip()
            goals = lines[1].replace('GOALS:', '').strip() if len(lines) > 1 else "Learn key concepts"
            
            return title, goals
        except Exception as e:
            print(f"Error generating chapter info: {e}")
            raise RuntimeError(f"Failed to generate chapter info for week {week_num}, chapter {chapter_num}: {e}")
    
    def estimate_lesson_time(self, lesson: Dict[str, str], week_lessons: List[List[Dict[str, str]]]) -> int:
        """Estimate time to complete a lesson in minutes."""
        # Simple estimation based on lesson title keywords
        title_lower = lesson['title'].lower()
        
        # Look up base time from dictionary
        base_time = BASE_TIME_ESTIMATES["default"]
        for keyword, time in BASE_TIME_ESTIMATES.items():
            if keyword != "default" and keyword in title_lower:
                base_time = time
                break
        
        return max(30, base_time)  # Minimum 30 minutes
    
    def adjust_times_to_30_hours(self, week_lessons: List[List[Dict[str, str]]], estimated_times: List[List[int]]) -> List[List[int]]:
        """Adjust lesson times so the week totals exactly 30 hours (1800 minutes)."""
        total_estimated = sum(sum(chapter_times) for chapter_times in estimated_times)
        target_minutes = 1800  # 30 hours
        
        if total_estimated == 0:
            return estimated_times
        
        # Calculate scaling factor
        scale_factor = target_minutes / total_estimated
        
        # Apply scaling and round to nearest 5 minutes
        adjusted_times = []
        for chapter_times in estimated_times:
            adjusted_chapter = []
            for time in chapter_times:
                adjusted_time = round(time * scale_factor / 5) * 5
                adjusted_time = max(15, adjusted_time)  # Minimum 15 minutes
                adjusted_chapter.append(adjusted_time)
            adjusted_times.append(adjusted_chapter)
        
        return adjusted_times
    
    def setup_worksheet_headers(self):
        """Set up the worksheet headers and formatting."""
        headers = [
            "Week", "Learning Goal of the week", "Chapter number", 
            "Chapter learning goals", "Chapter title", "", "Lesson title", "Time to complete", "Mastery Outcome"
        ]
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = [8, 50, 15, 40, 25, 5, 35, 15, 50]
        for col, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[get_column_letter(col)].width = width
    
    def add_week_data(self, week_num: int, week_lessons: List[List[Dict[str, str]]], week_learning_goal: str):
        """Add data for one week to the worksheet."""
        current_row = self.worksheet.max_row + 1
        
        # Estimate times for all lessons
        estimated_times = []
        for chapter_lessons in week_lessons:
            chapter_times = [self.estimate_lesson_time(lesson, week_lessons) for lesson in chapter_lessons]
            estimated_times.append(chapter_times)
        
        # Adjust times to total 30 hours
        adjusted_times = estimated_times
        
        # Add data for each chapter
        for chapter_num, (chapter_lessons, chapter_times) in enumerate(zip(week_lessons, adjusted_times), 1):
            chapter_title, chapter_goals = self.generate_chapter_info(chapter_lessons, week_num, chapter_num)
            chapter_start_row = current_row
            
            for lesson_num, (lesson, time) in enumerate(zip(chapter_lessons, chapter_times)):
                # Week column (merged for all lessons in the week)
                if chapter_num == 1 and lesson_num == 0:
                    week_cell = self.worksheet.cell(row=current_row, column=1, value=week_num)
                    week_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Learning goal column (merged for all lessons in the week)
                if chapter_num == 1 and lesson_num == 0:
                    goal_cell = self.worksheet.cell(row=current_row, column=2, value=week_learning_goal)
                    goal_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Chapter number (only set on first lesson of chapter)
                if lesson_num == 0:
                    chapter_cell = self.worksheet.cell(row=current_row, column=3, value=f"{week_num}.{chapter_num}")
                    chapter_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Chapter learning goals (only set on first lesson of chapter)
                if lesson_num == 0:
                    goals_cell = self.worksheet.cell(row=current_row, column=4, value=chapter_goals)
                    goals_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Chapter title (only set on first lesson of chapter)
                if lesson_num == 0:
                    title_cell = self.worksheet.cell(row=current_row, column=5, value=chapter_title)
                    title_cell.alignment = Alignment(vertical='top')
                
                # Empty column (column 6)
                
                # Lesson title
                lesson_cell = self.worksheet.cell(row=current_row, column=7, value=lesson['title'])
                lesson_cell.alignment = Alignment(vertical='top')
                
                # Time to complete
                time_cell = self.worksheet.cell(row=current_row, column=8, value=f"{time}")
                time_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Mastery Outcome
                outcome_cell = self.worksheet.cell(row=current_row, column=9, value=lesson['learning_outcomes'])
                outcome_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                current_row += 1
            
            # Merge cells for chapter columns (chapter number, goals, title)
            chapter_end_row = current_row - 1
            if chapter_start_row < chapter_end_row:
                self.worksheet.merge_cells(f'C{chapter_start_row}:C{chapter_end_row}')
                self.worksheet.merge_cells(f'D{chapter_start_row}:D{chapter_end_row}')
                self.worksheet.merge_cells(f'E{chapter_start_row}:E{chapter_end_row}')
        
        # Merge cells for week and learning goal columns
        start_row = current_row - sum(len(chapter) for chapter in week_lessons)
        end_row = current_row - 1
        
        # Merge week column
        if start_row < end_row:
            self.worksheet.merge_cells(f'A{start_row}:A{end_row}')
            self.worksheet.merge_cells(f'B{start_row}:B{end_row}')
    
    def save_syllabus_backup(self, lessons: List[Dict[str, str]], weeks_data: Dict[int, Tuple[str, List[List[Dict[str, str]]]]], num_practice_sessions: int, backup_file: str = 'output/syllabus_data.json'):
        """Save complete syllabus data to a JSON backup file.
        
        Args:
            lessons: Original list of lessons
            weeks_data: Dictionary mapping week_num to (learning_goal, week_lessons_with_practice)
            num_practice_sessions: Number of practice sessions per chapter
            backup_file: Path to the backup JSON file
        """
        # Ensure output directory exists
        os.makedirs(os.path.dirname(backup_file) if os.path.dirname(backup_file) else '.', exist_ok=True)
        
        # Organize all data for backup
        backup_data = {
            'lessons': lessons,
            'num_weeks': NUM_OF_WEEKS,
            'num_chapters_per_week': NUM_OF_CHAPTERS_PER_WEEK,
            'num_practice_sessions': num_practice_sessions,
            'weeks': {}
        }
        
        # Collect detailed week data
        for week_num, (week_learning_goal, week_lessons_with_practice) in weeks_data.items():
            # Estimate times for all lessons in this week
            estimated_times = []
            for chapter_lessons in week_lessons_with_practice:
                chapter_times = [self.estimate_lesson_time(lesson, week_lessons_with_practice) for lesson in chapter_lessons]
                estimated_times.append(chapter_times)
            
            # Collect week data for backup
            week_data = {
                'week_num': week_num,
                'learning_goal': week_learning_goal,
                'chapters': []
            }
            
            for chapter_num, chapter_lessons in enumerate(week_lessons_with_practice, 1):
                chapter_title, chapter_goals = self.generate_chapter_info(chapter_lessons, week_num, chapter_num)
                
                chapter_data = {
                    'chapter_num': chapter_num,
                    'title': chapter_title,
                    'learning_goals': chapter_goals,
                    'lessons': []
                }
                
                # Add lessons with times
                chapter_times = estimated_times[chapter_num - 1]
                for lesson, time in zip(chapter_lessons, chapter_times):
                    lesson_data = {
                        'title': lesson['title'],
                        'learning_outcomes': lesson['learning_outcomes'],
                        'time_minutes': time
                    }
                    chapter_data['lessons'].append(lesson_data)
                
                week_data['chapters'].append(chapter_data)
            
            backup_data['weeks'][week_num] = week_data
        
        # Save to JSON file
        with open(backup_file, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, indent=2, ensure_ascii=False)
        print(f"Backup saved to {backup_file}")
    
    def export_maestro_json(self, lessons: List[Dict[str, str]], weeks_data: Dict[int, Tuple[str, List[List[Dict[str, str]]]]], output_file: str):
        """Export syllabus data to Maestro JSON format using AI transformation."""
        # Build the lesson data organized by week
        weeks_lessons_text = ""
        for week_num in range(1, NUM_OF_WEEKS + 1):
            if week_num in weeks_data:
                week_learning_goal, week_lessons_with_practice = weeks_data[week_num]
                weeks_lessons_text += f"\n\nWEEK {week_num} - {week_learning_goal}\n"
                weeks_lessons_text += "Lessons:\n"
                
                # Flatten all lessons from all chapters in this week
                for chapter_lessons in week_lessons_with_practice:
                    for lesson in chapter_lessons:
                        is_practice = "practice session" in lesson['title'].lower()
                        marker = "[PRACTICE]" if is_practice else ""
                        weeks_lessons_text += f"  - {marker} {lesson['title']}: {lesson['learning_outcomes']}\n"
        
        # Build dynamic template structure based on NUM_OF_WEEKS
        template_units = []
        for i in range(1, NUM_OF_WEEKS + 1):
            template_units.append(f'''            {{
              "title": "Unit {i} Title Placeholder",
              "lessons": [
                {{
                  "title": "Lesson 1 Title Placeholder",
                  "masteryOutcomes": [
                    "Single mastery outcome placeholder"
                  ],
                  "teachingInstructions": ""
                }}
              ]
            }}''')
        
        units_template = ",\n".join(template_units)
        
        prompt = get_maestro_json_prompt(units_template, NUM_OF_WEEKS, weeks_lessons_text)
        
        try:
            client = self._get_openai_client()
            print(f"Generating Maestro JSON export...")
            
            # Use streaming for real-time feedback
            # Note: gpt-4 has 8K context, gpt-4-turbo has 128K
            # Adjust max_tokens based on available context after prompt
            stream = client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                **self._get_max_tokens_param(3600),  # Balanced for gpt-4; use gpt-4-turbo for larger outputs
                temperature=0.3,
                stream=True,
                stream_options={"include_usage": True}
            )
            
            # Collect chunks and track usage
            ai_content = ""
            usage_data = None
            for chunk in stream:
                if chunk.choices[0].delta.content is not None:
                    content_chunk = chunk.choices[0].delta.content
                    ai_content += content_chunk
                    # Show progress
                    if len(ai_content) % 200 == 0:
                        print(".", end="", flush=True)
                # Capture usage data from the final chunk
                if hasattr(chunk, 'usage') and chunk.usage is not None:
                    usage_data = chunk.usage
            
            print()  # New line after progress dots
            
            # Display cost information
            if usage_data:
                cost = self._calculate_cost(usage_data.prompt_tokens, usage_data.completion_tokens)
                print(f"📊 Tokens: {usage_data.prompt_tokens:,} in + {usage_data.completion_tokens:,} out = {usage_data.total_tokens:,} total | 💰 ${cost:.4f}")
            
            # Log the API call
            self._log_api_call(prompt, ai_content, "MAESTRO JSON EXPORT")
            
            # Clean up the response - extract JSON from markdown or plain text
            ai_content = ai_content.strip()
            
            # Try to extract JSON from markdown code blocks
            if '```json' in ai_content:
                # Extract content between ```json and ```
                start = ai_content.find('```json') + 7
                end = ai_content.find('```', start)
                if end != -1:
                    ai_content = ai_content[start:end].strip()
            elif '```' in ai_content:
                # Extract content between ``` and ```
                start = ai_content.find('```') + 3
                end = ai_content.find('```', start)
                if end != -1:
                    ai_content = ai_content[start:end].strip()
            
            # Try to find JSON object by looking for opening and closing braces
            if not ai_content.startswith('{'):
                # Find the first { in the content
                json_start = ai_content.find('{')
                if json_start != -1:
                    ai_content = ai_content[json_start:]
            
            # Check if response was likely truncated
            if not ai_content.endswith('}'):
                print("Warning: Response appears truncated.")
                print("TIP: Run with --model gpt-4-turbo for larger context window and better JSON generation.")
            
            # Parse and validate JSON
            try:
                maestro_data = json.loads(ai_content)
                print("Successfully parsed Maestro JSON")
            except json.JSONDecodeError as e:
                print(f"Warning: AI response is not valid JSON. Error: {e}")
                print(f"Response length: {len(ai_content)} characters")
                print("TIP: Run with --model gpt-4-turbo for better JSON generation and larger context.")
                print("Attempting to save raw response for debugging...")
                maestro_data = {
                    "error": "Invalid JSON from AI", 
                    "error_message": str(e),
                    "response_length": len(ai_content),
                    "raw_response": ai_content
                }
            
            # Determine output filename
            base_name = output_file.replace('.xlsx', '').replace('.xls', '')
            maestro_file = f"output/maestro_{base_name}.json"
            
            # Ensure output directory exists
            os.makedirs('output', exist_ok=True)
            
            # Save to file
            with open(maestro_file, 'w', encoding='utf-8') as f:
                json.dump(maestro_data, f, indent=2, ensure_ascii=False)
            
            print(f"Maestro JSON exported to {maestro_file}")
            
        except Exception as e:
            print(f"Error exporting Maestro JSON: {e}")
            print("Continuing with syllabus generation...")
    
    def generate_syllabus(self, lessons_file: str, output_file: str):
        """Generate the complete syllabus Excel file."""
        print("Reading lessons from file...")
        lessons = self.read_lessons_from_file(lessons_file)
        print(f"Found {len(lessons)} lessons")
        
        print("Distributing lessons across weeks...")
        distribution = self.distribute_lessons(lessons)

        print("Setting up worksheet...")
        self.setup_worksheet_headers()
        
        # Store minimal week data for backup
        weeks_data = {}
        
        print("Generating learning goals and creating syllabus...")
        for week_num in range(1, NUM_OF_WEEKS + 1):
            week_lessons = distribution[week_num]
            week_learning_goal, week_lessons_with_practice = self.generate_learning_goals(week_lessons, week_num, NUM_PRACTICE_SESSIONS_PER_CHAPTER)
            weeks_data[week_num] = (week_learning_goal, week_lessons_with_practice)
            self.add_week_data(week_num, week_lessons_with_practice, week_learning_goal)
            print(f"Week {week_num} completed")
        
        # Save all syllabus data to a backup json file
        if EXPORT_TO_LESSONS_JSON:
            self.save_syllabus_backup(lessons, weeks_data, NUM_PRACTICE_SESSIONS_PER_CHAPTER)
        
        # Export to Maestro JSON format
        if EXPORT_TO_MAESTRO_JSON:
            export_maestro_json(lessons, weeks_data, output_file)

        print(f"Saving syllabus to {output_file}...")
        self.workbook.save("output/" + output_file)
        print("Syllabus generated successfully!")
        
        # Log completion
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{self._get_timestamp()}] PROCESS COMPLETED\n")
            f.write("-" * 60 + "\n")
            f.write(f"Syllabus saved to: {output_file}\n")
            f.write(f"Total lessons processed: {len(lessons)}\n")
            f.write("=" * 80 + "\n\n")

def main():
    parser = argparse.ArgumentParser(description='Generate course syllabus from lesson titles')
    parser.add_argument('--lessons_file', help='Text file containing lesson titles (one per line)')
    parser.add_argument('-o', '--output', default='course_syllabus.xlsx', 
                       help='Output Excel file name (default: course_syllabus.xlsx)')
    parser.add_argument('--api-key', help='OpenAI API key (or set OPENAI_API_KEY environment variable)')
    parser.add_argument('--log-file', default='openai_log.txt',
                       help='Log file for OpenAI API calls (default: openai_log.txt)')
    parser.add_argument('--model', default=DEFAULT_MODEL,
                       help=f'OpenAI model to use (default: {DEFAULT_MODEL}, options: gpt-5.1, gpt-4o, gpt-4, gpt-4-turbo, gpt-3.5-turbo)')

    
    args = parser.parse_args()
    
    # Get API key
    api_key = args.api_key or os.getenv('OPENAI_API_KEY')
    if not api_key:
        print("Error: OpenAI API key required. Provide via --api-key argument or OPENAI_API_KEY environment variable.")
        sys.exit(1)
    
    # Generate syllabus
    generator = SyllabusGenerator(api_key, args.log_file, args.model)
    generator.generate_syllabus(args.lessons_file, args.output)

if __name__ == "__main__":
    main()
