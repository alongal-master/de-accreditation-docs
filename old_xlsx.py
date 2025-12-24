#!/usr/bin/env python3
"""
XLSX to Maestro JSON Converter

This script reads an existing syllabus XLSX file and converts it to Maestro JSON format.
The XLSX file should have the format: Week, Learning Goal, Chapter number, Chapter learning goals,
Chapter title, (empty), Lesson title, Time to complete, Mastery Outcome
"""

import argparse
import os
import sys
import json
from typing import List, Dict
import openai
from openpyxl import load_workbook
from dotenv import load_dotenv
from prompts import get_maestro_json_prompt

# Constants 
DIR = "output//"
#DEFAULT_MODEL = "gpt-4-turbo"
DEFAULT_MODEL = "gpt-4o"

# Pricing per 1M tokens (as of 2024, update as needed)
MODEL_PRICING = {
    "gpt-4o": {"input": 2.50, "output": 10.00},
    "gpt-4o-mini": {"input": 0.15, "output": 0.60},
    "gpt-4-turbo": {"input": 10.00, "output": 30.00},
    "gpt-4": {"input": 30.00, "output": 60.00},
    "gpt-3.5-turbo": {"input": 0.50, "output": 1.50},
}

# Load environment variables
load_dotenv()


class XLSXToMaestroConverter:
    def __init__(self, api_key: str, log_file: str = "maestro_log.txt", model: str = DEFAULT_MODEL):
        """Initialize the converter with OpenAI API key."""
        self.api_key = api_key
        self.log_file = log_file
        self.model = model
        self._init_log_file()
    
    def _init_log_file(self):
        """Initialize the log file with header."""
        with open(self.log_file, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("OpenAI API Log - Maestro JSON Converter\n")
            f.write("=" * 80 + "\n")
            f.write(f"Started at: {self._get_timestamp()}\n")
            f.write(f"Model: {self.model}\n")
            f.write("=" * 80 + "\n\n")
    
    def _get_timestamp(self):
        """Get current timestamp for logging."""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def _log_api_call(self, prompt: str, response: str, call_type: str):
        """Log API call details to file."""
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{self._get_timestamp()}] {call_type}\n")
            f.write("-" * 60 + "\n")
            f.write("PROMPT:\n")
            f.write(prompt)
            f.write("\n\n")
            f.write("RESPONSE:\n")
            f.write(response)
            f.write("\n\n")
            f.write("=" * 80 + "\n\n")
    
    def _get_openai_client(self):
        """Get OpenAI client."""
        return openai.OpenAI(api_key=self.api_key)
    
    def _calculate_cost(self, input_tokens: int, output_tokens: int) -> float:
        """Calculate the cost of an API call based on token usage."""
        if self.model not in MODEL_PRICING:
            print(f"Warning: Pricing not available for model {self.model}")
            return 0.0
        
        pricing = MODEL_PRICING[self.model]
        input_cost = (input_tokens / 1_000_000) * pricing["input"]
        output_cost = (output_tokens / 1_000_000) * pricing["output"]
        total_cost = input_cost + output_cost
        return total_cost
    
    def read_xlsx(self, xlsx_file: str) -> Dict:
        """Read and parse the XLSX file into structured data."""
        try:
            print(f"Reading XLSX file: {xlsx_file}")
            workbook = load_workbook(xlsx_file)
            worksheet = workbook.active
            
            # Parse the data
            weeks_data = {}
            current_week = None
            current_chapter = None
            
            # Skip header row (row 1)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                week_num = row[0]  # Column A: Week
                week_goal = row[1]  # Column B: Learning Goal of the week
                chapter_num = row[2]  # Column C: Chapter number
                chapter_goals = row[3]  # Column D: Chapter learning goals
                chapter_title = row[4]  # Column E: Chapter title
                # row[5] is empty column
                lesson_title = row[6]  # Column G: Lesson title
                time_to_complete = row[7]  # Column H: Time to complete
                mastery_outcome = row[8] if len(row) > 8 else ""  # Column I: Mastery Outcome
                
                # Skip empty rows
                if not lesson_title:
                    continue
                
                # Initialize week if new
                if week_num and week_num != current_week:
                    # Convert to int to handle float/string from Excel
                    try:
                        current_week = int(float(week_num))
                    except (ValueError, TypeError):
                        print(f"Warning: Skipping row with invalid week number: {week_num}")
                        continue
                    
                    if current_week not in weeks_data:
                        weeks_data[current_week] = {
                            'week_num': current_week,
                            'learning_goal': week_goal or "",
                            'chapters': []
                        }
                
                # Initialize chapter if new
                if chapter_num and chapter_num != current_chapter:
                    # Convert to int to handle float/string from Excel
                    try:
                        current_chapter = int(float(chapter_num))
                    except (ValueError, TypeError):
                        print(f"Warning: Skipping row with invalid chapter number: {chapter_num}")
                        continue
                    
                    weeks_data[current_week]['chapters'].append({
                        'chapter_num': current_chapter,
                        'title': chapter_title or "",
                        'learning_goals': chapter_goals or "",
                        'lessons': []
                    })
                
                # Add lesson to current chapter
                if current_week and weeks_data[current_week]['chapters']:
                    lesson = {
                        'title': lesson_title,
                        'learning_outcomes': mastery_outcome or "",
                        'time_minutes': time_to_complete
                    }
                    weeks_data[current_week]['chapters'][-1]['lessons'].append(lesson)
            
            print(f"Successfully parsed {len(weeks_data)} weeks")
            return weeks_data
            
        except FileNotFoundError:
            print(f"Error: File '{xlsx_file}' not found.")
            sys.exit(1)
        except Exception as e:
            print(f"Error reading XLSX file: {e}")
            sys.exit(1)
    
    def generate_maestro_json(self, weeks_data: Dict, output_file: str):
        """Generate Maestro JSON from structured weeks data using AI."""
        # Build the lesson data organized by week
        weeks_lessons_text = ""
        num_weeks = len(weeks_data)
        
        for week_num in sorted(weeks_data.keys()):
            week_info = weeks_data[week_num]
            week_learning_goal = week_info['learning_goal']
            weeks_lessons_text += f"\n\nWEEK {week_num} - {week_learning_goal}\n"
            weeks_lessons_text += "Lessons:\n"
            
            # Flatten all lessons from all chapters in this week
            for chapter in week_info['chapters']:
                for lesson in chapter['lessons']:
                    is_practice = "practice session" in lesson['title'].lower()
                    marker = "[PRACTICE]" if is_practice else ""
                    weeks_lessons_text += f"  - {marker} {lesson['title']}: {lesson['learning_outcomes']}\n"
        
        # Build dynamic template structure based on number of weeks
        template_units = []
        for i in range(1, num_weeks + 1):
            template_units.append(f'''            {{
              "title": "Unit {i} Title Placeholder",
              "lessons": [
                {{
                  "title": "Lesson 1 Title Placeholder",
                  "masteryOutcomes": [
                    "Single mastery outcome placeholder"
                  ],
                  "teachingInstructions": ""
                }}
              ]
            }}''')
        
        units_template = ",\n".join(template_units)
        
        # Get prompt from prompts.py
        prompt = get_maestro_json_prompt(units_template, num_weeks, weeks_lessons_text)
        
        try:
            client = self._get_openai_client()
            print(f"Generating Maestro JSON export with {self.model}...")
            
            # Use streaming for real-time feedback
            stream = client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=16000,  # Increased for large courses with many lessons
                temperature=0.3,
                stream=True,
                stream_options={"include_usage": True}
            )
            
            # Collect chunks and track usage
            ai_content = ""
            usage_data = None
            for chunk in stream:
                # Skip chunks without choices
                if not chunk.choices or len(chunk.choices) == 0:
                    continue
                    
                if chunk.choices[0].delta.content is not None:
                    content_chunk = chunk.choices[0].delta.content
                    ai_content += content_chunk
                    # Show progress
                    if len(ai_content) % 200 == 0:
                        print(".", end="", flush=True)
                # Capture usage data from the final chunk
                if hasattr(chunk, 'usage') and chunk.usage is not None:
                    usage_data = chunk.usage
            
            print()  # New line after progress dots
            
            # Display cost information
            if usage_data:
                input_tokens = usage_data.prompt_tokens
                output_tokens = usage_data.completion_tokens
                total_tokens = usage_data.total_tokens
                cost = self._calculate_cost(input_tokens, output_tokens)
                print(f"📊 Token Usage: {input_tokens:,} input + {output_tokens:,} output = {total_tokens:,} total")
                print(f"💰 Estimated Cost: ${cost:.4f}")
            else:
                print("⚠️  Usage data not available")
            
            # Log the API call
            self._log_api_call(prompt, ai_content, "MAESTRO JSON EXPORT")
            
            # Clean up the response - extract JSON from markdown or plain text
            ai_content = ai_content.strip()
            
            # Try to extract JSON from markdown code blocks
            if '```json' in ai_content:
                start = ai_content.find('```json') + 7
                end = ai_content.find('```', start)
                if end != -1:
                    ai_content = ai_content[start:end].strip()
            elif '```' in ai_content:
                start = ai_content.find('```') + 3
                end = ai_content.find('```', start)
                if end != -1:
                    ai_content = ai_content[start:end].strip()
            
            # Try to find JSON object by looking for opening and closing braces
            if not ai_content.startswith('{'):
                json_start = ai_content.find('{')
                if json_start != -1:
                    ai_content = ai_content[json_start:]
            
            # Check if response was likely truncated
            if not ai_content.endswith('}'):
                print("Warning: Response appears truncated.")
            
            # Parse and validate JSON
            try:
                maestro_data = json.loads(ai_content)
                print("Successfully parsed Maestro JSON")
            except json.JSONDecodeError as e:
                print(f"Warning: AI response is not valid JSON. Error: {e}")
                print(f"Response length: {len(ai_content)} characters")
                print("Attempting to save raw response for debugging...")
                maestro_data = {
                    "error": "Invalid JSON from AI", 
                    "error_message": str(e),
                    "response_length": len(ai_content),
                    "raw_response": ai_content
                }
            
            # Save to file
            os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(maestro_data, f, indent=2, ensure_ascii=False)
            
            print(f"Maestro JSON saved to {output_file}")
            
        except Exception as e:
            print(f"Error generating Maestro JSON: {e}")
            sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description='Convert syllabus XLSX to Maestro JSON format')
    parser.add_argument('xlsx_file', help='Input XLSX file with syllabus data')
    parser.add_argument('-o', '--output', help='Output JSON file name (default: maestro_[input_name].json)')
    parser.add_argument('--api-key', help='OpenAI API key (or set OPENAI_API_KEY environment variable)')
    parser.add_argument('--log-file', default='maestro_log.txt',
                       help='Log file for OpenAI API calls (default: maestro_log.txt)')
    parser.add_argument('--model', default=DEFAULT_MODEL,
                       help=f'OpenAI model to use (default: {DEFAULT_MODEL}, options: gpt-4, gpt-4-turbo)')
    
    args = parser.parse_args()
    
    # Get API key
    api_key = args.api_key or os.getenv('OPENAI_API_KEY')
    if not api_key:
        print("Error: OpenAI API key required. Provide via --api-key argument or OPENAI_API_KEY environment variable.")
        sys.exit(1)
    
    # Determine output filename
    if args.output:
        output_file = f"{DIR}{args.output}"
    else:
        # Create default output filename based on input
        base_name = os.path.basename(args.xlsx_file).replace('.xlsx', '').replace('.xls', '')
        output_file = f"{DIR}maestro_{base_name}.json"
    
    # Convert XLSX to Maestro JSON
    converter = XLSXToMaestroConverter(api_key, args.log_file, args.model)
    weeks_data = converter.read_xlsx(args.xlsx_file)
    converter.generate_maestro_json(weeks_data, output_file)
    
    print("Conversion completed successfully!")


if __name__ == "__main__":
    main()

