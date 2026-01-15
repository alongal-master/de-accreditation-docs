#!/usr/bin/env python3
"""
XLSX to Maestro JSON Converter (Non-LLM Version)

This script reads an existing syllabus XLSX file and converts it to Maestro JSON format.
The XLSX file should have the format: Week, Learning Goal, Chapter number, Chapter learning goals,
Chapter title, (empty), Lesson title, Time to complete, Mastery Outcome
"""

import argparse
import os
import sys
import json
import re
from typing import Dict
from datetime import datetime
from openpyxl import load_workbook

# Constants
DIR = "output_json//"

# Course metadata defaults
DEFAULT_COURSE_TITLE = ""
DEFAULT_COURSE_DESCRIPTION = ""
DEFAULT_TRACK_TYPE = "Programming"
DEFAULT_VERSION = 1
DEFAULT_CREDITS = 0
DEFAULT_DISPLAY_ID = ""
DEFAULT_LABEL = ""
DEFAULT_IS_PUBLISHED = False

# Lesson filtering
SKIP_LESSON_KEYWORDS = ["Closing session","Opening session","Sync Session", "Weekly Review"]

# Practice session configuration
PRACTICE_SESSION_KEYWORDS = ["Practice Lesson", "Practice Session"]
PRACTICE_TEACHING_INSTRUCTIONS = """This is a coding challenge lesson. In this lesson, do not introduce new topics; it is about solving an exercise using previously learned concepts only.
Guide the student step by step toward the solution without writing the full answer or final code for them.
Break the problem into small, manageable steps and ask the student to implement each step before moving on.
When the student is stuck, give progressive hints instead of solutions.
You may provide example inputs/outputs, edge cases, and clarification to help the student reason about correctness.
Review the student's work by pointing out what's correct and what needs improvement, then suggest the next step.
Exercise theme is to be chosen based on the lesson title.
Exercise goals and desired output: """

# Theory Practice session configuration
THEORY_PRACTICE_KEYWORD = "Theory Practice Lesson"
THEORY_PRACTICE_TEACHING_INSTRUCTIONS = """Create a challenge-based lesson that is grounded in the material covered in the **current unit until this lesson, where the student demonstrates the knowledge they have learned in a fun and engaging way. Avoid coding in this lesson.
The lesson should include between eight and ten interactive and dynamic rounds between the student and the Maestro. 
The challenge may include different types of questions or learning experiences, such as varied question formats, interactive tasks, MCQ, identification or matching questions, etc.
You are free to choose any structure or format that best supports an engaging challenge experience.
During the challenge itself, there is no need to provide feedback or corrections, the focus should remain entirely on the challenge experience.
After all challenge rounds are completed, provide a short summary that offers encouraging feedback, highlights areas for professional improvement and refinement, and points out the student's strengths as demonstrated through their responses during the challenge.
Ensure the lesson remains aligned with the topics that were taught and is appropriate for the student's level. Don't code in this lesson.
Lesson goals: """

# Lesson defaults
DEFAULT_REQUIRED_PLUGINS = ["code-editor"]
DEFAULT_TEACHING_INSTRUCTIONS = ""

# Lesson title prefixes - maps starting characters/strings to prefix text
# Example: {"X": "Prefix text "} will add "Prefix text " to all lessons starting with "X"
LESSON_TITLE_PREFIXES = {
    "Theory Practice Lesson": "📚 "
}
# Add all practice session keywords with the same prefix
for keyword in PRACTICE_SESSION_KEYWORDS:
    LESSON_TITLE_PREFIXES[keyword] = "⚙ "


class XLSXToMaestroConverter:
    def __init__(self):
        """Initialize the converter."""
        pass
    
    def read_xlsx(self, xlsx_file: str) -> Dict:
        """Read and parse the XLSX file into structured data."""
        try:
            print(f"Reading XLSX file: {xlsx_file}")
            workbook = load_workbook(xlsx_file)
            worksheet = workbook.active
            
            # Parse the data
            weeks_data = {}
            current_week = None
            current_chapter = None
            
            # Skip header row (row 1)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                week_num = row[0]  # Column A: Week
                week_goal = row[1]  # Column B: Learning Goal of the week
                chapter_num = row[2]  # Column C: Chapter number
                chapter_goals = row[3]  # Column D: Chapter learning goals
                chapter_title = row[4]  # Column E: Chapter title
                # row[5] is empty column
                lesson_title = row[6]  # Column G: Lesson title
                time_to_complete = row[7]  # Column H: Time to complete
                mastery_outcome = row[8] if len(row) > 8 else ""  # Column I: Mastery Outcome
                
                # Skip empty rows
                if not lesson_title:
                    continue
                
                # Initialize week if new
                if week_num and week_num != current_week:
                    # Convert to int to handle float/string from Excel
                    try:
                        current_week = int(float(week_num))
                    except (ValueError, TypeError):
                        print(f"Warning: Skipping row with invalid week number: {week_num}")
                        continue
                    
                    if current_week not in weeks_data:
                        weeks_data[current_week] = {
                            'week_num': current_week,
                            'learning_goal': week_goal or "",
                            'chapters': []
                        }
                
                # Initialize chapter if new
                if chapter_num and chapter_num != current_chapter:
                    # Convert to int to handle float/string from Excel
                    try:
                        current_chapter = int(float(chapter_num))
                    except (ValueError, TypeError):
                        print(f"Warning: Skipping row with invalid chapter number: {chapter_num}")
                        continue
                    
                    weeks_data[current_week]['chapters'].append({
                        'chapter_num': current_chapter,
                        'title': chapter_title or "",
                        'learning_goals': chapter_goals or "",
                        'lessons': []
                    })
                
                # Add lesson to current chapter
                if current_week and weeks_data[current_week]['chapters']:
                    lesson = {
                        'title': lesson_title,
                        'learning_outcomes': mastery_outcome or "",
                        'time_minutes': time_to_complete
                    }
                    weeks_data[current_week]['chapters'][-1]['lessons'].append(lesson)
            
            print(f"Successfully parsed {len(weeks_data)} weeks")
            return weeks_data
            
        except FileNotFoundError:
            print(f"Error: File '{xlsx_file}' not found.")
            sys.exit(1)
        except Exception as e:
            print(f"Error reading XLSX file: {e}")
            sys.exit(1)
    
    def _should_skip_lesson(self, lesson_title: str) -> bool:
        """Check if a lesson should be skipped based on keywords."""
        if not lesson_title:
            return True
        lesson_lower = lesson_title.lower()
        return any(keyword.lower() in lesson_lower for keyword in SKIP_LESSON_KEYWORDS)
    
    def _is_practice_session(self, lesson_title: str) -> bool:
        """Check if a lesson is a practice session (but not theory practice)."""
        if not lesson_title:
            return False
        title_lower = lesson_title.lower()
        # Make sure it's not a theory practice lesson
        if THEORY_PRACTICE_KEYWORD.lower() in title_lower:
            return False
        # Check against list of practice session keywords
        return any(keyword.lower() in title_lower for keyword in PRACTICE_SESSION_KEYWORDS)
    
    def _is_theory_practice_session(self, lesson_title: str) -> bool:
        """Check if a lesson is a theory practice session."""
        if not lesson_title:
            return False
        return THEORY_PRACTICE_KEYWORD.lower() in lesson_title.lower()
    
    def _split_mastery_outcomes(self, outcomes: str) -> list:
        """Split mastery outcomes by semicolons and clean up."""
        if not outcomes:
            return []
        # Split by semicolon and strip whitespace
        split_outcomes = [outcome.strip() for outcome in outcomes.split(';')]
        # Filter out empty strings
        return [outcome for outcome in split_outcomes if outcome]
    
    def _extract_week_title(self, learning_goal: str) -> str:
        """Extract just the title part from learning goal (before colon)."""
        if not learning_goal:
            return ""
        # Split by colon and take the first part, strip whitespace
        title = learning_goal.split(':', 1)[0].strip()
        return title
    
    def _extract_week_description(self, learning_goal: str) -> str:
        """Extract the description part from learning goal (after colon)."""
        if not learning_goal:
            return ""
        # Split by colon and take the second part if it exists, strip whitespace
        parts = learning_goal.split(':', 1)
        if len(parts) > 1:
            return parts[1].strip()
        return ""
    
    def _apply_lesson_prefix(self, lesson_title: str) -> str:
        """Apply prefix to lesson title if it starts with a configured prefix key."""
        if not lesson_title or not LESSON_TITLE_PREFIXES:
            return lesson_title
        
        for prefix_key, prefix_text in LESSON_TITLE_PREFIXES.items():
            if lesson_title.startswith(prefix_key):
                return f"{prefix_text}{lesson_title}"
        
        return lesson_title
    
    def _strip_leading_number(self, title: str) -> str:
        """Strip leading numbers like '4. ' or '12. ' from title."""
        if not title:
            return title
        # Match patterns like "4. ", "12. ", "4) ", "4 - ", "4: " at the start
        cleaned = re.sub(r'^\d+[\.\)\-:\s]+\s*', '', title)
        return cleaned if cleaned else title
    
    def _normalize_dashes(self, text: str) -> str:
        """Replace em dashes (—), en dashes (–), and other dash variants with regular hyphens (-)."""
        if not text:
            return text
        # Replace em dash (—), en dash (–), and other unicode dashes with regular hyphen
        return re.sub(r'[—–−‐‑‒―]', '-', text)
    
    def generate_maestro_json(self, weeks_data: Dict, output_file: str, xlsx_filename: str = ""):
        """Generate Maestro JSON from structured weeks data."""
        print("Generating Maestro JSON export...")
        
        num_weeks = len(weeks_data)
        
        # Extract course title from XLSX filename (base name without extension)
        if xlsx_filename:
            course_title = os.path.basename(xlsx_filename)
            # Remove .xlsx or .xls extension
            course_title = course_title.replace('.xlsx', '').replace('.xls', '')
        else:
            course_title = DEFAULT_COURSE_TITLE
        
        # Build units from weeks
        units = []
        for week_num in sorted(weeks_data.keys()):
            week_info = weeks_data[week_num]
            week_learning_goal = week_info['learning_goal']
            
            # Extract title and description from learning goal (format: "Title: Description")
            week_title = self._extract_week_title(week_learning_goal)
            unit_title = week_title if week_title else f"WEEK {week_num}"
            unit_description = self._extract_week_description(week_learning_goal)
            
            # Flatten all lessons from all chapters in this week
            lessons = []
            for chapter in week_info['chapters']:
                for lesson in chapter['lessons']:
                    original_lesson_title = lesson['title']
                    
                    # Skip lessons that match skip keywords
                    if self._should_skip_lesson(original_lesson_title):
                        continue
                    
                    # Check for practice sessions on ORIGINAL title (before any processing)
                    is_theory_practice = self._is_theory_practice_session(original_lesson_title)
                    is_practice = self._is_practice_session(original_lesson_title)
                    
                    # Strip leading numbers (e.g., "4. Data-Driven AI" -> "Data-Driven AI")
                    lesson_title = self._strip_leading_number(original_lesson_title)
                    
                    # Normalize dashes (em dash → hyphen)
                    lesson_title = self._normalize_dashes(lesson_title)
                    
                    # Apply prefix if lesson title starts with configured prefix key
                    lesson_title = self._apply_lesson_prefix(lesson_title)
                    
                    # Split mastery outcomes by semicolons and normalize dashes
                    mastery_outcomes = [self._normalize_dashes(o) for o in self._split_mastery_outcomes(lesson['learning_outcomes'])]
                    
                    # Determine teaching instructions and mastery outcomes
                    if is_theory_practice:
                        # Theory practice lessons: put everything in teaching instructions, no mastery outcomes
                        # Strip "Output:" part from learning outcomes
                        outcomes_text = lesson['learning_outcomes']
                        if 'Output:' in outcomes_text:
                            outcomes_text = outcomes_text.split('Output:')[0].strip()
                        elif 'output:' in outcomes_text:
                            outcomes_text = outcomes_text.split('output:')[0].strip()
                        teaching_instructions = THEORY_PRACTICE_TEACHING_INSTRUCTIONS + self._normalize_dashes(outcomes_text)
                        mastery_outcomes = []
                    elif is_practice:
                        # Practice lessons: put everything in teaching instructions, no mastery outcomes
                        teaching_instructions = PRACTICE_TEACHING_INSTRUCTIONS + self._normalize_dashes(lesson['learning_outcomes'])
                        mastery_outcomes = []
                    else:
                        teaching_instructions = DEFAULT_TEACHING_INSTRUCTIONS
                    
                    # Build lesson object
                    lesson_obj = {
                        "title": lesson_title,
                        "masteryOutcomes": mastery_outcomes,
                        "teachingInstructions": teaching_instructions,
                        "requiredPlugins": DEFAULT_REQUIRED_PLUGINS.copy()
                    }
                    
                    lessons.append(lesson_obj)
            
            # Create unit object
            unit_obj = {
                "title": unit_title,
                "description": unit_description,
                "lessons": lessons
            }
            
            units.append(unit_obj)
        
        # Build the course JSON structure (just the course object, not the full track)
        maestro_data = {
            "description": DEFAULT_COURSE_DESCRIPTION,
            "credits": DEFAULT_CREDITS,
            "teachingInstructions": DEFAULT_TEACHING_INSTRUCTIONS,
            "units": units
        }
        
        # Save to file
        os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(maestro_data, f, indent=2, ensure_ascii=False)
        
        print(f"Maestro JSON saved to {output_file}")
        print("Conversion completed successfully!")
    
    def generate_maestro_json_from_units(self, units_data: Dict, output_file: str, xlsx_filename: str = ""):
        """Generate Maestro JSON from unit-grouped data (lessons grouped by original_unit).
        
        Args:
            units_data: OrderedDict mapping unit_name to {'unit_name': str, 'lessons': [...]}
            output_file: Output file path
            xlsx_filename: Original filename (used to derive course title)
        """
        print("Generating Maestro JSON export from units...")
        
        # Extract course title from XLSX filename (base name without extension)
        if xlsx_filename:
            course_title = os.path.basename(xlsx_filename)
            course_title = course_title.replace('.xlsx', '').replace('.xls', '')
        else:
            course_title = DEFAULT_COURSE_TITLE
        
        # Build units from the units_data
        units = []
        for unit_name, unit_info in units_data.items():
            # Extract title and description from unit_name (format: "Title: Description")
            unit_title = self._extract_week_title(unit_name)
            unit_description = self._extract_week_description(unit_name)
            
            # Process lessons for this unit
            lessons = []
            for lesson in unit_info['lessons']:
                original_lesson_title = lesson['title']
                
                # Skip lessons that match skip keywords
                if self._should_skip_lesson(original_lesson_title):
                    continue
                
                # Check for practice sessions on ORIGINAL title (before any processing)
                is_theory_practice = self._is_theory_practice_session(original_lesson_title)
                is_practice = self._is_practice_session(original_lesson_title)
                
                # Strip leading numbers
                lesson_title = self._strip_leading_number(original_lesson_title)
                
                # Normalize dashes
                lesson_title = self._normalize_dashes(lesson_title)
                
                # Apply prefix if lesson title starts with configured prefix key
                lesson_title = self._apply_lesson_prefix(lesson_title)
                
                # Split mastery outcomes and normalize dashes
                mastery_outcomes = [self._normalize_dashes(o) for o in self._split_mastery_outcomes(lesson['learning_outcomes'])]
                
                # Determine teaching instructions and mastery outcomes
                if is_theory_practice:
                    # Theory practice lessons: strip Output: part
                    outcomes_text = lesson['learning_outcomes']
                    if 'Output:' in outcomes_text:
                        outcomes_text = outcomes_text.split('Output:')[0].strip()
                    elif 'output:' in outcomes_text:
                        outcomes_text = outcomes_text.split('output:')[0].strip()
                    teaching_instructions = THEORY_PRACTICE_TEACHING_INSTRUCTIONS + self._normalize_dashes(outcomes_text)
                    mastery_outcomes = []
                elif is_practice:
                    teaching_instructions = PRACTICE_TEACHING_INSTRUCTIONS + self._normalize_dashes(lesson['learning_outcomes'])
                    mastery_outcomes = []
                else:
                    teaching_instructions = DEFAULT_TEACHING_INSTRUCTIONS
                
                # Build lesson object
                lesson_obj = {
                    "title": lesson_title,
                    "masteryOutcomes": mastery_outcomes,
                    "teachingInstructions": teaching_instructions,
                    "requiredPlugins": DEFAULT_REQUIRED_PLUGINS.copy()
                }
                
                lessons.append(lesson_obj)
            
            # Create unit object
            unit_obj = {
                "title": unit_title,
                "description": unit_description,
                "lessons": lessons
            }
            
            units.append(unit_obj)
        
        # Build the course JSON structure (just the course object, not the full track)
        maestro_data = {
            "description": DEFAULT_COURSE_DESCRIPTION,
            "credits": DEFAULT_CREDITS,
            "teachingInstructions": DEFAULT_TEACHING_INSTRUCTIONS,
            "units": units
        }
        
        # Save to file
        os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(maestro_data, f, indent=2, ensure_ascii=False)
        
        print(f"Maestro JSON saved to {output_file}")
        print(f"Generated {len(units)} units from original unit structure")


def main():
    parser = argparse.ArgumentParser(description='Convert syllabus XLSX to Maestro JSON format (Non-LLM version)')
    parser.add_argument('xlsx_file', help='Input XLSX file with syllabus data')
    parser.add_argument('-o', '--output', help='Output JSON file name (default: maestro_[input_name].json)')
    
    args = parser.parse_args()
    
    # Determine output filename
    if args.output:
        output_file = f"{DIR}{args.output}"
    else:
        # Create default output filename based on input
        base_name = os.path.basename(args.xlsx_file).replace('.xlsx', '').replace('.xls', '')
        output_file = f"{DIR}maestro_{base_name}.json"
    
    # Convert XLSX to Maestro JSON
    converter = XLSXToMaestroConverter()
    weeks_data = converter.read_xlsx(args.xlsx_file)
    converter.generate_maestro_json(weeks_data, output_file, args.xlsx_file)


if __name__ == "__main__":
    main()

