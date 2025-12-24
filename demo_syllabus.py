#!/usr/bin/env python3
"""
Demo script that creates a sample syllabus without requiring OpenAI API.
This shows the structure and formatting of the generated Excel file.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_demo_syllabus():
    """Create a demo syllabus with sample data."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Course Syllabus"
    
    # Sample data structure
    demo_data = {
        1: {
            "learning_goal": "Python Fundamentals: Master basic Python syntax, variables, and control structures to build a solid programming foundation.",
            "chapters": [
                {
                    "number": "1.1",
                    "title": "Getting Started",
                    "goals": "Learn Python setup, basic syntax, and first programs",
                    "lessons": [
                        ("Introduction to Python Programming", 30),
                        ("Setting up Python Environment", 45),
                        ("Variables and Data Types", 60)
                    ]
                },
                {
                    "number": "1.2", 
                    "title": "Basic Operations",
                    "goals": "Master arithmetic operations and string manipulation",
                    "lessons": [
                        ("Basic Input and Output", 45),
                        ("Arithmetic Operations", 30),
                        ("String Manipulation", 60)
                    ]
                },
                {
                    "number": "1.3",
                    "title": "Control Flow",
                    "goals": "Understand conditional statements and loops",
                    "lessons": [
                        ("Control Structures - If Statements", 75),
                        ("Control Structures - Loops", 90)
                    ]
                }
            ]
        },
        2: {
            "learning_goal": "Data Structures & Functions: Explore Python's built-in data structures and learn to create reusable functions for efficient programming.",
            "chapters": [
                {
                    "number": "2.1",
                    "title": "Data Structures",
                    "goals": "Master lists, dictionaries, and sets for data organization",
                    "lessons": [
                        ("Lists and List Operations", 90),
                        ("Dictionaries and Sets", 75),
                        ("Data Structure Practice", 60)
                    ]
                },
                {
                    "number": "2.2",
                    "title": "Functions",
                    "goals": "Create and use functions for code reusability",
                    "lessons": [
                        ("Functions - Definition and Usage", 90),
                        ("Function Parameters and Return Values", 75),
                        ("Modules and Packages", 60)
                    ]
                },
                {
                    "number": "2.3",
                    "title": "File Handling",
                    "goals": "Read from and write to files for data persistence",
                    "lessons": [
                        ("File Handling - Reading Files", 60),
                        ("File Handling - Writing Files", 60),
                        ("Error Handling and Exceptions", 75)
                    ]
                }
            ]
        },
        3: {
            "learning_goal": "Object-Oriented Programming: Design and implement classes, understand inheritance, and build modular, maintainable code.",
            "chapters": [
                {
                    "number": "3.1",
                    "title": "OOP Basics",
                    "goals": "Understand classes, objects, and encapsulation",
                    "lessons": [
                        ("Object-Oriented Programming Basics", 90),
                        ("Classes and Objects", 90),
                        ("Inheritance and Polymorphism", 90)
                    ]
                },
                {
                    "number": "3.2",
                    "title": "Advanced Topics",
                    "goals": "Explore advanced programming concepts and libraries",
                    "lessons": [
                        ("Working with Libraries", 60),
                        ("Data Analysis with Pandas", 90),
                        ("Data Visualization with Matplotlib", 90)
                    ]
                },
                {
                    "number": "3.3",
                    "title": "Web Integration",
                    "goals": "Connect Python to web services and APIs",
                    "lessons": [
                        ("Web Scraping Introduction", 75),
                        ("Working with APIs", 75),
                        ("Database Connections", 90)
                    ]
                }
            ]
        },
        4: {
            "learning_goal": "Project Development & Best Practices: Apply all learned concepts in a comprehensive project while mastering testing, debugging, and optimization techniques.",
            "chapters": [
                {
                    "number": "4.1",
                    "title": "Testing & Debugging",
                    "goals": "Learn professional development practices",
                    "lessons": [
                        ("Testing Your Code", 90),
                        ("Debugging Techniques", 75),
                        ("Code Optimization", 60)
                    ]
                },
                {
                    "number": "4.2",
                    "title": "Project Planning",
                    "goals": "Plan and structure a complete Python project",
                    "lessons": [
                        ("Project Planning and Structure", 90),
                        ("Final Project Development", 120),
                        ("Project Presentation", 60)
                    ]
                }
            ]
        }
    }
    
    # Set up headers
    headers = [
        "Week", "Learning Goal of the week", "Chapter number", 
        "Chapter learning goals", "Chapter title", "", "Lesson title", "Time to complete"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    column_widths = [8, 50, 15, 40, 25, 5, 35, 15]
    for col, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = width
    
    # Add data
    current_row = 2
    
    for week_num, week_data in demo_data.items():
        week_start_row = current_row
        
        for chapter in week_data["chapters"]:
            for lesson_title, lesson_time in chapter["lessons"]:
                # Week column
                if current_row == week_start_row:
                    week_cell = worksheet.cell(row=current_row, column=1, value=week_num)
                    week_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Learning goal column
                if current_row == week_start_row:
                    goal_cell = worksheet.cell(row=current_row, column=2, value=week_data["learning_goal"])
                    goal_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Chapter number
                chapter_cell = worksheet.cell(row=current_row, column=3, value=chapter["number"])
                chapter_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Chapter learning goals
                goals_cell = worksheet.cell(row=current_row, column=4, value=chapter["goals"])
                goals_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Chapter title
                title_cell = worksheet.cell(row=current_row, column=5, value=chapter["title"])
                title_cell.alignment = Alignment(vertical='top')
                
                # Empty column (column 6)
                
                # Lesson title
                lesson_cell = worksheet.cell(row=current_row, column=7, value=lesson_title)
                lesson_cell.alignment = Alignment(vertical='top')
                
                # Time to complete
                time_cell = worksheet.cell(row=current_row, column=8, value=f"{lesson_time} min")
                time_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += 1
        
        # Merge cells for week and learning goal columns
        week_end_row = current_row - 1
        if week_start_row < week_end_row:
            worksheet.merge_cells(f'A{week_start_row}:A{week_end_row}')
            worksheet.merge_cells(f'B{week_start_row}:B{week_end_row}')
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, current_row):
        for col in range(1, 9):
            worksheet.cell(row=row, column=col).border = thin_border
    
    # Save the file
    output_file = "demo_syllabus.xlsx"
    workbook.save(output_file)
    print(f"Demo syllabus created: {output_file}")
    print(f"Total lessons: {sum(len(chapter['lessons']) for week in demo_data.values() for chapter in week['chapters'])}")
    print(f"Total weeks: {len(demo_data)}")
    
    # Calculate total time per week
    for week_num, week_data in demo_data.items():
        total_minutes = sum(lesson_time for chapter in week_data["chapters"] for _, lesson_time in chapter["lessons"])
        hours = total_minutes / 60
        print(f"Week {week_num}: {total_minutes} minutes ({hours:.1f} hours)")

if __name__ == "__main__":
    create_demo_syllabus()
